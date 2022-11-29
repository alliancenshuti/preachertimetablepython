from openpyxl import load_workbook

wb = load_workbook(filename='preachersDb.xlsx')
sheet_ranges = wb['preachers']
rows = sheet_ranges.max_row
columns = sheet_ranges.max_column


def fetch_data():
    preachers = []
    for i in range(3, rows - 2):
        preacher = {"name": sheet_ranges.cell(row=i, column=1).value, "inconvenient": sheet_ranges.cell(row=i, column=3).value.split(",")}
        preachers.append(preacher)
    return preachers
